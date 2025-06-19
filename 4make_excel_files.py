import glob
import json
import os
from typing import Dict, List, Any, Set, Tuple, Optional

import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

def read_json_files(folder_path: str) -> List[Dict]:
    """
    Read all JSON files from the specified directory in order.

    Args:
        folder_path: Path to directory containing JSON files

    Returns:
        List of dictionaries containing all JSON data
    """
    all_data = []

    if not os.path.exists(folder_path):
        print(f"Error: Directory {folder_path} does not exist")
        return all_data

    # Get all JSON files and sort them
    json_files = sorted(glob.glob(os.path.join(folder_path, "*.json")))

    if not json_files:
        print(f"No JSON files found in {folder_path}")
        return all_data

    print(f"Found {len(json_files)} JSON files in {folder_path}")

    for file_path in json_files:
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                all_data.append(data)
                print(f"Successfully read: {os.path.basename(file_path)}")
        except Exception as e:
            print(f"Error reading {file_path}: {e}")

    return all_data


def extract_special_fields(record: Dict) -> Dict[str, str]:
    """
    Extract and format special fields: 'Связанные сообщения' and 'Предметы финансовой аренды (лизинга)'.

    Args:
        record: Single record from JSON data

    Returns:
        Dictionary with formatted special fields
    """
    special_fields = {}

    # Handle "Связанные сообщения"
    if 'Сообщение' in record and 'Связанные сообщения' in record['Сообщение']:
        related_messages = record['Сообщение']['Связанные сообщения']
        formatted_messages = []

        for key, value in related_messages.items():
            formatted_messages.append(f'{key}: "{value}"')

        special_fields['Связанные сообщения'] = '\n'.join(formatted_messages)

    # Handle "Предметы финансовой аренды (лизинга)"
    if 'Сообщение' in record and 'Предметы финансовой аренды (лизинга)' in record['Сообщение']:
        items = record['Сообщение']['Предметы финансовой аренды (лизинга)']

        identifiers = []
        classifiers = []
        descriptions = []

        for key, item in items.items():
            identifiers.append(f"{key}. {item.get('Идентификатор', 'нет данных')}")
            classifiers.append(f"{key}. {item.get('Классификатор', 'нет данных')}")
            descriptions.append(f"{key}. {item.get('Описание', 'нет данных')}")

        special_fields['Идентификатор'] = ' \n'.join(identifiers)
        special_fields['Классификатор'] = ' \n'.join(classifiers)
        special_fields['Описание'] = ' \n'.join(descriptions)

    return special_fields


def parse_lessor_info(lessor_str: str) -> Tuple[Optional[str], Optional[str]]:
    """
    Parse lessor string to extract INN and OGRN values.

    Args:
        lessor_str: Raw string from "Лизингодатели" field

    Returns:
        Tuple containing (INN, OGRN) with None values if not found
    """
    if not isinstance(lessor_str, str):
        return None, None

    # Split into parts and clean empty strings
    parts = [part.strip() for part in lessor_str.split('\n') if part.strip()]

    # Try to find INN/OGRN using predictable patterns
    inn, ogrn = None, None

    try:
        for i, part in enumerate(parts):
            if part == "ИНН" and i + 1 < len(parts):
                inn = parts[i + 1]
            elif part == "ОГРН" and i + 1 < len(parts):
                ogrn = parts[i + 1]

        # Validate extracted values (simple digit checks)
        if inn and not inn.isdigit():
            inn = None
        if ogrn and not ogrn.isdigit():
            ogrn = None
    except Exception:
        pass

    return inn, ogrn


def flatten_record(record: Dict, parent_key: str = '', separator: str = ' ') -> Dict[str, Any]:
    """
    Flatten nested dictionary structure.

    Args:
        record: Dictionary to flatten
        parent_key: Parent key for nested structure
        separator: Separator for nested keys

    Returns:
        Flattened dictionary
    """
    items = []

    for key, value in record.items():
        new_key = f"{key} ({parent_key})" if parent_key else key

        if isinstance(value, dict):
            # Skip special nested objects that are handled separately
            if key in ['Связанные сообщения', 'Предметы финансовой аренды (лизинга)', 'ЗАГОЛОВОК']:
                continue
            items.extend(flatten_record(value, new_key, separator).items())
        else:
            items.append((new_key, value))

    return dict(items)


def get_all_columns(all_records: List[Dict]) -> Set[str]:
    """
    Get all unique column names from all records.

    Args:
        all_records: List of all flattened records

    Returns:
        Set of all unique column names
    """
    all_columns = set()

    for record in all_records:
        all_columns.update(record.keys())

    return all_columns


def process_single_record(url: str, record_data: Dict) -> Dict[str, Any]:
    """
    Process a single record and extract all data.

    Args:
        url: URL key for the record
        record_data: The record data

    Returns:
        Processed record as dictionary
    """
    # Start with the URL
    processed_record = {'url': url}

    # Extract special fields first
    special_fields = extract_special_fields(record_data)
    processed_record.update(special_fields)

    # Extract header fields
    if 'ЗАГОЛОВОК' in record_data:
        header_data = record_data['ЗАГОЛОВОК']
        processed_record['Основной заголовок'] = header_data.get('Основной заголовок', '')
        processed_record['Подзаголовок'] = header_data.get('Подзаголовок', '')
    else:
        processed_record['Основной заголовок'] = ''
        processed_record['Подзаголовок'] = ''

    # Extract lessor info if available
    lessor_field = None
    if 'Сообщение' in record_data and 'Лизингодатели' in record_data['Сообщение']:
        lessor_field = record_data['Сообщение']['Лизингодатели']

    inn, ogrn = parse_lessor_info(lessor_field) if lessor_field else (None, None)
    processed_record['ИНН Лизингодателя'] = inn
    processed_record['ОГРН Лизингодателя'] = ogrn

    # Flatten the regular fields
    flattened = flatten_record(record_data)
    processed_record.update(flattened)

    return processed_record


def convert_to_excel(json_data_list: List[Dict], output_file: str) -> None:
    """
    Convert JSON data to Excel format with clickable hyperlinks
    """
    # Step 1: Process all records
    processed_records = process_records(json_data_list)
    if not processed_records:
        return

    # Step 2: Create and organize DataFrame
    df = create_dataframe(processed_records)

    # Step 3: Save with hyperlink processing
    save_with_hyperlinks(df, output_file)


def process_records(json_data_list: List[Dict]) -> List[Dict]:
    """Process all JSON records into structured dictionaries"""
    processed_records = []
    for json_data in json_data_list:
        for url, record_data in json_data.items():
            processed_record = process_single_record(url, record_data)
            processed_records.append(processed_record)

    if not processed_records:
        print("No records to process")
    return processed_records


def create_dataframe(processed_records: List[Dict]) -> pd.DataFrame:
    """Create organized DataFrame with proper column ordering"""
    df = pd.DataFrame(processed_records)

    # Define column groups and ordering
    special_columns = [
        'url',
        'Основной заголовок',
        'Подзаголовок',
        'ИНН Лизингодателя',
        'ОГРН Лизингодателя',
        'Идентификатор',
        'Классификатор',
        'Описание',
        'Связанные сообщения'
    ]

    # Filter existing columns and order them
    existing_special_cols = [col for col in special_columns if col in df.columns]
    other_columns = sorted([col for col in df.columns if col not in special_columns])

    return df.reindex(columns=existing_special_cols + other_columns)


def save_with_hyperlinks(df: pd.DataFrame, output_file: str) -> None:
    """Save DataFrame to Excel with proper hyperlink formatting"""
    import tempfile

    # Create a temporary file with .xlsx extension
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmpfile:
        temp_path = tmpfile.name
        print(f"Using temporary file: {temp_path}")

    try:
        # Step 1: Save initial version to temp file
        df.to_excel(temp_path, index=False, engine='openpyxl')
        print(f"Temporary file saved: {temp_path}")

        # Step 2: Add hyperlinks and re-save
        add_hyperlinks(temp_path, output_file)
        print(f"Saved final file with hyperlinks: {output_file}")

        # Cleanup
        if os.path.exists(temp_path):
            os.remove(temp_path)

    except Exception as e:
        print(f"Error processing hyperlinks: {e}")
        # Fallback: try to copy without hyperlinks
        try:
            df.to_excel(output_file, index=False)
            print(f"Saved without hyperlinks as: {output_file}")
        except Exception as e2:
            print(f"Final fallback failed too: {e2}")

def add_hyperlinks(input_path: str, output_path: str) -> None:
    """Add Excel hyperlink formatting to URL column"""
    from openpyxl import load_workbook

    wb = load_workbook(input_path)
    ws = wb.active

    # Find URL column
    url_col_idx = find_column_index(ws, 'url')
    if not url_col_idx:
        print("URL column not found - saving without hyperlinks")
        wb.save(output_path)
        return

    # Apply hyperlink formatting
    blue_font = Font(color="0000FF", underline="single")
    for row in range(2, ws.max_row + 1):  # Skip header row
        cell = ws.cell(row=row, column=url_col_idx)
        url = cell.value
        if valid_url(url):
            cell.hyperlink = url
            cell.font = blue_font

    # Adjust column widths for readability
    auto_adjust_columns(ws)

    wb.save(output_path)
    print(f"Saved final workbook with hyperlinks: {output_path}")


# Helper functions
def find_column_index(worksheet, column_name: str) -> int:
    """Find column index by header name"""
    for idx, col in enumerate(worksheet.iter_cols(max_row=1), 1):
        if col[0].value == column_name:
            return idx
    return None


def valid_url(url: Any) -> bool:
    """Validate URL format"""
    return isinstance(url, str) and url.startswith('http')


def auto_adjust_columns(worksheet, max_width: int = 100) -> None:
    """Auto-adjust column widths with maximum limit"""
    from openpyxl.utils import get_column_letter

    for column in worksheet.columns:
        col_letter = get_column_letter(column[0].column)
        max_length = 0

        for cell in column:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        adjusted_width = min(max_length + 2, max_width)
        worksheet.column_dimensions[col_letter].width = adjusted_width


def main():
    """
    Main function to process JSON files and create Excel output.
    """
    # Define the directory path
    directory_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "3raw_contents")

    # Output file name changed to "output.xlsx"
    output_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output.xlsx")

    try:
        # Read all JSON files from the directory
        print(f"Reading JSON files from directory: {directory_path}")
        json_data_list = read_json_files(directory_path)

        if not json_data_list:
            print("No JSON data found. Please check your directory path and ensure it contains JSON files.")
            return

        # Convert to Excel
        print("Converting to Excel...")
        convert_to_excel(json_data_list, output_file)

        print("Process completed successfully!")

    except Exception as e:
        print(f"Error during processing: {e}")


if __name__ == "__main__":
    main()
